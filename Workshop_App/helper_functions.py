import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, Fill
from openpyxl.styles.borders import Border
from PyQt5.QtWidgets import QFileDialog, QMainWindow
from json import load
from gui_window import GuiWindow
from workshop_tool import WorkshopsTool


def generate_workshop_info(ui: GuiWindow, ws: WorkshopsTool) -> None:
    """Output the desired content based on selected options."""

    ui.textOutputField.clear()
    ws.set_search_phrase(ui.lineEditPhrase.text())

    update_searched_workshops(ui, ws)

    button_check: bool = check_button_options(ui)

    text_to_display: str = get_workshop_display_text(ui, ws, button_check)

    ui.textOutputField.clear()
    ui.textOutputField.insertPlainText(text_to_display)


def update_searched_workshops(ui: GuiWindow, ws: WorkshopsTool) -> list:
    """Return the correct workshops based on selected options."""

    if ui.lineEditWorkshopID.text() != "":
        workshops = ws.get_matching_workshops_by_id(ui.lineEditWorkshopID.text())
    elif ui.checkBoxUseDate.isChecked():
        starting_date: tuple = ui.calendarWidget_StartDate.selectedDate().getDate()
        ending_date: tuple = ui.calendarWidget_EndDate.selectedDate().getDate()        
        workshops  = ws.get_matching_workshops_by_date_range(starting_date, ending_date)
    else:
        workshops  = ws.get_matching_workshops()

    return workshops


def check_button_options(ui: GuiWindow) -> bool:
    """Return true if any button is checked."""

    return (
        ui.checkBoxWsID.isChecked()
        or ui.checkBoxWsStartDate.isChecked()
        or ui.checkBoxWsPartNumbers.isChecked()
        or ui.checkBoxWsName.isChecked()
        or ui.checkBoxWsURL.isChecked()
    )


def update_database(main_window: QMainWindow, ws: WorkshopsTool, ui: GuiWindow) -> None:
    """Attempts to handle connecting to the website and database."""

    ui.textOutputField.clear()

    try:
        ws.setup_workshop_information()
        ui.textOutputField.insertPlainText(get_welcome_text())
    except ConnectionError:
        ui.textOutputField.insertPlainText(get_welcome_text_for_offline())
    except TypeError:
        ui.textOutputField.insertPlainText(get_server_error_text())
    except FileNotFoundError as e:
        ui.textOutputField.insertPlainText(get_missing_file_text())

    main_window.repaint()


def get_workshop_display_text(ui: GuiWindow, ws: WorkshopsTool, button_check: bool) -> str:
    display_text = list()
    display_text.append(f"Number of matching workshops: {ws.get_number_of_workshops()}\n\n")
    display_text.append(f"Total Signed Up: {ws.get_number_of_participants()}\n\n")

    if button_check:
        display_text = setup_workshop_information_text(ui, display_text, ws)

    display_text.append(f"All emails for these workshops:\n\n{ws.get_emails()}")
    
    return "".join(display_text)


def get_welcome_text() -> str:
    """Get text that first appears in output window."""    

    welcome_text: list = [
        "Your database has been updated!",
        "",
        "This program will allow you to seach current workshops using a phrase, date range, or Session ID.",
        'Type a phrase that you would like to search in the "Phrase:" field.',
        'Leave the "Phrase:" field blank to get all current workshops.',
        "The Session ID search will take priority over phrase and date range search.",
    ] 

    return "\n".join(welcome_text)


def get_welcome_text_for_offline() -> str:
    """Get text that first appears in output window if offline."""

    offline_text: list = [
        "Cannot connect to server...",
        "You are in offline mode.",
        "You can access workshops from the last time your last successful connection.",
        "You can try to update your database when you have established an internet connection.",
    ]

    return "\n".join(offline_text)


def get_server_error_text() -> str:
    """Return server error message."""

    server_error_text: str = "Something went wrong when connecting to server...\nTry again later."

    return server_error_text


def get_missing_file_text() -> str:
    """Return missing file message. """

    missing_file_text: str = 'Missing "connection_info.json". Cannot update database.'

    return missing_file_text


def setup_workshop_information_text(ui: GuiWindow, display_text: str, ws: WorkshopsTool) -> str:
    """Prepare the output of the workshops based on selected information."""

    for workshop in ws.get_most_recent_search_results():
        text = list()
        if ui.checkBoxWsID.isChecked():
            text.append(f"{workshop['workshop_id']}")
        if ui.checkBoxWsStartDate.isChecked():
            text.append(f"{workshop['workshop_start_date_and_time']}")
        if ui.checkBoxWsPartNumbers.isChecked():
            text.append(f"{workshop['workshop_signed_up']}/{workshop['workshop_participant_capacity']}")
        if ui.checkBoxWsName.isChecked():
            text.append(f"{workshop['workshop_name']}")
        if ui.checkBoxWsURL.isChecked():
            text.append(f"\n   Url: {workshop['workshop_url']}")

        display_text.append(" - ".join(text))
        display_text.append("\n")

        if (
            ui.checkBoxNames.isChecked()
            or ui.checkBoxEmails.isChecked()
            or ui.checkBoxSchools.isChecked()
        ):

            display_text.append("   Contact Information:\n")

            for participant_info in workshop["workshop_participant_info_list"]:
                text = list()
                if ui.checkBoxNames.isChecked():
                    text.append(f'{participant_info["name"]}')
                if ui.checkBoxEmails.isChecked():
                    text.append(f'{participant_info["email"]}')
                if ui.checkBoxSchools.isChecked():
                    text.append(f'{participant_info["school"]}')
                display_text.append("    + ")
                display_text.append(" - ".join(text))
                display_text.append("\n")

        display_text.append("\n")

    return display_text

def export_workshops_info(ui: GuiWindow, ws: WorkshopsTool) -> None:
    """Exports the searched workshop information to an .xlsx file."""

    ws.set_search_phrase(ui.lineEditPhrase.text())

    workshops: list = ws.get_most_recent_search_results()

    workbook = Workbook()
    workbook["Sheet"].title = "Workshops"

    workshops_sheet = workbook.active
    attendance_sheet = workbook.create_sheet("Attendance")

    attendance_sheet.merge_cells("A1:B1")
    attendance_sheet.merge_cells("C1:E1")
    attendance_sheet.merge_cells("A2:B2")
    attendance_sheet.merge_cells("C2:E2")
    attendance_sheet["A1"] = "Workshop Name:"
    attendance_sheet["C1"] = "<NAME_HERE>"
    attendance_sheet["A2"] = "Workshop Dates:"
    attendance_sheet["C2"] = "<DATES_HERE>"
    attendance_sheet.append([])

    co_op_abbreviations = list()
    co_op_session_location: dict = get_co_op_session_locations()

    workshop_rows = list()

    for workshop in workshops:
        row: list = build_row_for_workshop(co_op_session_location, workshop)
        workshop_rows.append(row)

    workshop_rows.sort()

    for row in workshop_rows:
        workshops_sheet.append(row[:8])
        
        sheet = workbook.create_sheet(row[0])        
        sheet.append(row[:8])

        attendance_sheet.append([row[0], row[1], row[7]])
        current_row = attendance_sheet._current_row
        attendance_sheet.merge_cells(f"C{current_row}:E{current_row}")
        co_op_abbreviations.append(row[0])
        attendance_sheet.append(["Name", "Email", "District", "Hours", "Dates Attended"])

        if len(row[8]) > 0:
            for partipant in row[8]:
                sheet.append([partipant["name"], partipant["email"], partipant["school"]])

                attendance_row: list = [partipant["name"], partipant["email"], partipant["school"]]
                attendance_sheet.append(attendance_row)

        attendance_sheet.append([])

    attendance_sheet_last_row: int = attendance_sheet._current_row
    format_sheet(attendance_sheet, attendance_sheet_last_row, co_op_abbreviations)

    save_file_info: str = QFileDialog().getSaveFileName(None, directory="workshop_info.xlsx", filter="Excel files (*.xlsx)")[0]

    # Only save file if the user provided a file name and didn't cancel.
    if save_file_info != "":
        workbook.save(filename=save_file_info)


def get_potential_name(ws: list) -> str:
    return ws[0]["workshop_name"]


def build_row_for_workshop(co_op_session_location: dict, workshop: dict) -> list:
    """Build out the contents of one spread sheet row entry. """
    
    location: list = workshop["workshop_location"].split(" - ")[0]

    row = [
        co_op_session_location[location]["abbr"],
        workshop["workshop_id"],
        workshop["workshop_name"],
        workshop["workshop_start_date_and_time"],
        workshop["workshop_signed_up"],
        workshop["workshop_participant_capacity"],
        co_op_session_location[location]["pd_doc_text"],
        workshop["workshop_url"],
        workshop["workshop_participant_info_list"]
    ]

    return row


def format_sheet(worksheet, last_row: int, coops: list) -> None:    
    """Formats excel spreadsheet for export."""
    
    worksheet.column_dimensions["A"].width = 40
    worksheet.column_dimensions["B"].width = 40
    worksheet.column_dimensions["C"].width = 40
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 40
    
    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    dark_grey_fill = PatternFill(fill_type="solid", start_color="222222", end_color="222222") 
    grey_fill = PatternFill(fill_type="solid", start_color="DDDDDD", end_color="DDDDDD")   

    title_header_font = Font(size=16, bold=True)
    title_header_alignment = Alignment(horizontal="right")
    title_content_font = Font(size=16, bold=True, color="AEEBAE")
    title_content_alignment = Alignment(horizontal="center")

    worksheet["A1"].font = title_header_font 
    worksheet["A2"].font = title_header_font
    worksheet["A1"].alignment = title_header_alignment 
    worksheet["A2"].alignment = title_header_alignment    
    worksheet["A1"].border = Border(left=thick, top=thick, right=thick)
    worksheet["A2"].border = Border(left=thick, bottom=thick, right=thick)
    worksheet["C1"].font = title_content_font 
    worksheet["C2"].font = title_content_font
    worksheet["C1"].fill = dark_grey_fill 
    worksheet["C2"].fill = dark_grey_fill
    worksheet["C1"].alignment = title_content_alignment 
    worksheet["C2"].alignment = title_content_alignment
    worksheet["C1"].border = Border(left=thick, top=thick, right=thick)
    worksheet["C2"].border = Border(left=thick, bottom=thick, right=thick)

    columns: str = "ABCDE"
    row: int = 4
    coop_info_font = Font(size=12, bold=True)
    coop_info_fill = PatternFill(fill_type="solid", start_color="AEEBAE", end_color="AEEBAE")
    coop_info_border = Border(top=thick, bottom=thin)
    content_font = Font(size=12)
    right_aligned = Alignment(shrink_to_fit=True, horizontal="right")
    left_aligned = Alignment(shrink_to_fit=True, horizontal="left")

    while row <= last_row:
        current_cell = worksheet[f"A{row}"]
        if current_cell.value in coops:
            for column in columns[:3]:
                current_cell = worksheet[f"{column}{row}"]
                current_cell.font = coop_info_font
                current_cell.fill = coop_info_fill
                current_cell.border = coop_info_border
                current_cell.alignment = left_aligned
            current_cell.value = f'=HYPERLINK("{current_cell.value}")'
        elif current_cell.value == "Name":
            for column in columns:
                current_cell = worksheet[f"{column}{row}"]
                current_cell.font = content_font
                current_cell.fill = grey_fill
                if column not in columns[:3]:
                    current_cell.alignment = right_aligned
                else:
                    current_cell.alignment = left_aligned
        elif current_cell.value != None: 
            for column in columns:
                current_cell = worksheet[f"{column}{row}"]           
                current_cell.font = content_font
                if column not in columns[:3]:
                    current_cell.alignment = right_aligned
                else:
                    current_cell.alignment = left_aligned
        else:
            for column in columns:
                current_cell = worksheet[f"{column}{row}"]              

        row += 1      


def get_co_op_session_locations() -> dict:
    """Load and get the Session Location information from the co_op_names.json file."""

    with open("co_op_names.json", "r") as f:
        return load(f)

    
if __name__ == '__main__':
    print('This is a module...')