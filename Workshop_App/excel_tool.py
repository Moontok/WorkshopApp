from PyQt5.QtWidgets import QFileDialog
from workshop_tool import WorkshopsTool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, numbers
from openpyxl.styles.borders import Border
from gui_window import GuiWindow
from spread_sheet_base_tool import SpreadSheetBaseTool

class ExcelTool(SpreadSheetBaseTool):

    def __init__(self):
        self.co_op_abbreviations = list()

        self.line: dict = {
            "thick":Side(border_style="thick", color="000000"),
            "thin":Side(border_style="thin", color="000000")
        }
        self.fill: dict = {
            "dark_grey":PatternFill(fill_type="solid", start_color="222222", end_color="222222"), 
            "grey":PatternFill(fill_type="solid", start_color="DDDDDD", end_color="DDDDDD"),
            "light_green":PatternFill(fill_type="solid", start_color="AEEBAE", end_color="AEEBAE")
        }
        self.align: dict = {            
            "right":Alignment(horizontal="right"),
            "left":Alignment(horizontal="left"),
            "fill":Alignment(horizontal="fill"),
            "center":Alignment(horizontal="center")
        }

    def export_workshops_info(self, ui: GuiWindow, ws: WorkshopsTool) -> None:
        """Exports the searched workshop information to an .xlsx file."""

        ws.set_search_phrase(ui.lineEditPhrase.text())

        workshops: list = ws.get_most_recent_search_results()

        workbook = Workbook()
        workbook["Sheet"].title = "Workshops"

        workshops_sheet = workbook.active

        workshops_sheet.append(["Workshops At A Glance"])
        workshops_sheet.append([])

        attendance_sheet = workbook.create_sheet("Attendance")
        attendance_sheet["A1"] = "Workshop Name:"
        attendance_sheet["C1"] = workshops[0]["workshop_name"]
        attendance_sheet["A2"] = "Workshop Dates:"
        attendance_sheet["C2"] = self.format_dates(workshops[0])
        attendance_sheet.append([])

        workshop_rows = list()

        for workshop in workshops:
            row: list = self.build_row_for_workshop(ws.get_co_op_info(), workshop)
            workshop_rows.append(row)

        workshop_rows.sort()

        for row in workshop_rows:
            workshops_sheet.append(row[:8])
            
            sheet = workbook.create_sheet(row[0])        
            sheet.append(row[:3])
            sheet.append(["Dates:", ", ".join(row[11])])
            sheet.append(["Credit:", row[12]])
            sheet.append(["Fee:", row[13]])
            sheet.append(["Description:", row[9]])
            sheet.append(["Session Link:", row[7]])
            sheet.append(["Signed Up"])

            attendance_sheet.append([row[0], row[1], row[7]])
            attendance_sheet.append(["Name", "Email", "District", "Hours", "Dates Attended"])

            if len(row[8]) > 0:
                for partipant in row[8]:
                    sheet.append([partipant["name"], "", partipant["email"], partipant["school"]])

                    attendance_row: list = [partipant["name"], partipant["email"], partipant["school"]]
                    attendance_sheet.append(attendance_row)

            attendance_sheet.append([])
            self.co_op_abbreviations.append(row[0])            
            self.format_generated_ws_sheet(sheet)
        
        workshops_sheet[f"A{workshops_sheet._current_row + 2}"] = "Total:"
        workshops_sheet[f"B{workshops_sheet._current_row}"] = len(workshop_rows)
        workshops_sheet[f"D{workshops_sheet._current_row}"] = "Signed Up:"
        workshops_sheet[f"E{workshops_sheet._current_row}"] = f"=SUM(E3:E{workshops_sheet._current_row - 2})"
        
        self.format_workshops_sheet(workshops_sheet)
        self.format_attendance_sheet(attendance_sheet)

        save_file_info: str = QFileDialog().getSaveFileName(None, directory="workshop_info.xlsx", filter="Excel files (*.xlsx)")[0]

        # Only save file if the user provided a file name and didn't cancel.
        if save_file_info != "":
            workbook.save(filename=save_file_info)

    
    def format_workshops_sheet(self, worksheet) -> None:
        """Formats excel workshops sheet."""

        worksheet.merge_cells("A1:H1")
        worksheet["A1"].font = Font(size=18, bold=True)
        worksheet["A1"].fill = self.fill["light_green"]

        last_row: int = worksheet._current_row

        worksheet[f"D{last_row}"].font = Font(bold=True)
        worksheet[f"E{last_row}"].font = Font(bold=True)
        worksheet[f"A{last_row}"].font = Font(bold=True)
        worksheet[f"B{last_row}"].font = Font(bold=True)

        worksheet.column_dimensions["A"].width = 15
        worksheet.column_dimensions["C"].width = 70
        worksheet.column_dimensions["D"].width = 18
        worksheet.column_dimensions["G"].width = 45
        worksheet.column_dimensions["H"].width = 70

        for row in range(3, last_row - 1):
            worksheet[f"H{row}"].value = f'=HYPERLINK("{worksheet[f"H{row}"].value}")'


    def format_generated_ws_sheet(self, worksheet) -> None:
        """General format for each Co-op sheet."""
        
        worksheet["A1"].font = Font(size=18, bold=True)
        worksheet["A1"].fill = self.fill["light_green"]
        worksheet["B1"].font = Font(size=18, bold=True)
        worksheet["B1"].fill = self.fill["light_green"]

        worksheet.merge_cells("C1:D1")
        worksheet["C1"].font = Font(size=18, bold=True)
        worksheet["C1"].fill = self.fill["light_green"]

        worksheet.merge_cells("B2:D2")
        worksheet.merge_cells("B3:D3")
        worksheet.merge_cells("B4:D4")

        worksheet.merge_cells("B5:D5")        
        worksheet["A5"].alignment = Alignment(vertical="top")
        worksheet["B5"].alignment = Alignment(vertical="top", wrap_text=True)

        worksheet.merge_cells("B6:D6")
        worksheet["B6"].value = f'=HYPERLINK("{worksheet["B6"].value}")'

        worksheet.merge_cells("A7:D7")
        worksheet["A7"].font = Font(bold=True)
        worksheet["A7"].fill = self.fill["grey"]

        worksheet.column_dimensions["A"].width = 15
        worksheet.column_dimensions["B"].width = 15
        worksheet.column_dimensions["C"].width = 50
        worksheet.column_dimensions["D"].width = 50


        for row in range(8, worksheet._current_row + 1):
            worksheet.merge_cells(f"A{row}:B{row}")
            worksheet[f"D{row}"].alignment = self.align["fill"]


    def format_attendance_sheet(self, worksheet) -> None:    
        """Formats excel attendance sheet."""        

        worksheet.merge_cells("A1:B1")
        worksheet.merge_cells("C1:E1")
        worksheet.merge_cells("A2:B2")
        worksheet.merge_cells("C2:E2")

        worksheet.column_dimensions["A"].width = 40
        worksheet.column_dimensions["B"].width = 40
        worksheet.column_dimensions["C"].width = 40
        worksheet.column_dimensions["D"].width = 20
        worksheet.column_dimensions["E"].width = 40
        
         
        title_header_font = Font(size=16, bold=True)
        title_content_font = Font(size=16, bold=True, color="AEEBAE")

        worksheet["A1"].font = title_header_font 
        worksheet["A2"].font = title_header_font
        worksheet["A1"].alignment = self.align["right"]
        worksheet["A2"].alignment = self.align["center"]        
        worksheet["A1"].border = Border(left=self.line["thick"], top=self.line["thick"], right=self.line["thick"])
        worksheet["A2"].border = Border(left=self.line["thick"], bottom=self.line["thick"], right=self.line["thick"])
        worksheet["B1"].border = Border(top=self.line["thick"])
        worksheet["B2"].border = Border(bottom=self.line["thick"])

        worksheet["C1"].font = title_content_font 
        worksheet["C2"].font = title_content_font
        worksheet["C1"].fill = self.fill["dark_grey"] 
        worksheet["C2"].fill = self.fill["dark_grey"]
        worksheet["C1"].alignment = self.align["center"] 
        worksheet["C2"].alignment = self.align["center"]
        worksheet["C1"].border = Border(left=self.line["thick"], top=self.line["thick"], right=self.line["thick"])
        worksheet["C2"].border = Border(left=self.line["thick"], bottom=self.line["thick"], right=self.line["thick"])
        worksheet["D1"].border = Border(top=self.line["thick"])
        worksheet["E1"].border = Border(top=self.line["thick"], right=self.line["thick"])
        worksheet["D2"].border = Border(bottom=self.line["thick"])
        worksheet["E2"].border = Border(bottom=self.line["thick"], right=self.line["thick"])

        columns: str = "ABCDE"
        coop_info_font = Font(size=12, bold=True)
        coop_info_border = Border(
            top=self.line["thick"], 
            bottom=self.line["thin"]
        )
        content_font = Font(size=12)

        for row in range(4, worksheet._current_row + 1):
            current_cell = worksheet[f"A{row}"]
            if current_cell.value in self.co_op_abbreviations:
                for column in columns:
                    current_cell = worksheet[f"{column}{row}"]
                    current_cell.font = coop_info_font
                    current_cell.fill = self.fill["light_green"]
                    current_cell.border = coop_info_border
                    current_cell.alignment = self.align["right"]                
                worksheet.merge_cells(f"C{row}:E{row}")
                worksheet[f"C{row}"] = f'=HYPERLINK("{worksheet[f"C{row}"].value}")'
            elif current_cell.value == "Name":
                for column in columns:
                    current_cell = worksheet[f"{column}{row}"]
                    current_cell.font = content_font
                    current_cell.fill = self.fill["grey"]
                    if column not in columns[:3]:
                        current_cell.alignment = self.align["right"]
                    else:
                        current_cell.alignment = self.align["left"]
            elif current_cell.value != None: 
                for column in columns:
                    current_cell = worksheet[f"{column}{row}"]           
                    current_cell.font = content_font
                    if column not in columns[:3]:
                        current_cell.alignment = self.align["right"]
                    else:
                        current_cell.alignment = self.align["fill"]
            else:
                for column in columns:
                    current_cell = worksheet[f"{column}{row}"]              


if __name__ == "__main__":
    print("This is a module...")
